import tkinter as tk        # importing tkinter
from tkinter import ttk
import openpyxl

def toggle_mode():
    if mode_selector.instate(['selected']):
        style.theme_use('forest-light')

    else:
        style.theme_use('forest-dark')



def insert_row():
    name = name_entry.get()
    age = int(age_entry.get())
    subscription_status = status_entry.get()
    employement_status = "Employed" if bool.get() else "Unemployed"
    print(name, age, subscription_status, employement_status)

    path = "/Users/zarnabkhan/Documents/MSc-BDS_QMUL/Self Learning Projects/TKINTER EXCEL APP/people.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    row_values = [name, age, subscription_status, employement_status]
    sheet.append(row_values)
    workbook.save(path)

    data_view.insert('', tk.END, values = row_values)

    name_entry.delete(0, "end")
    name_entry.insert(0, "Name")
    age_entry.delete(0, "end")
    age_entry.insert(0, "Age")
    status_entry.set(combo_list[0])
    button.state(["!selected"])



root = tk.Tk()      # root widget 

style = ttk.Style(root)     # apply theme 

root.tk.call("source", "forest-light.tcl")      # custom light theme
root.tk.call("source", "forest-dark.tcl")       # cutom dark theme
style.theme_use("forest-dark")     # setting up the default them

combo_list = ["Subscribed", "Not Subscribed", "Other"]

frame = ttk.Frame(root)
frame.pack()        # framme remains in center

widgets_frame = ttk.LabelFrame(frame, text = "Insert Row")
widgets_frame.grid(row = 0, column=0, padx = 20, pady = 10)

name_entry = ttk.Entry(widgets_frame)
name_entry.insert(0, "Name")
name_entry.bind("<FocusIn>", lambda e: name_entry.delete('0', 'end'))       # focus in mean when we click on the entry it removes the entry
name_entry.grid(row = 0, column=0, padx = 5, pady = (0,5), sticky= "ew")        #ew = east west

age_entry = ttk.Spinbox(widgets_frame, from_= 18,to_ = 100)
age_entry.insert(0, "Age")
age_entry.grid(row = 1, column=0, padx = 5, pady = (0,5), sticky= "ew") 

status_entry = ttk.Combobox(widgets_frame, values= combo_list)
status_entry.current(0)
status_entry.grid(row = 2, column=0,padx = 5, pady = (0,5), sticky= "ew") 

bool = tk.BooleanVar()
employement_entry = ttk.Checkbutton(widgets_frame, text ="Employed", variable = bool)
employement_entry.grid(row = 3, column=0, padx = 5, pady = (0,5), sticky= "nsew")

button = ttk.Button(widgets_frame, text ="Insert" ,command= insert_row)
button.grid(row = 4, column=0,padx = 5, pady = (0,5), sticky= "nsew")

separator = ttk.Separator(widgets_frame)
separator.grid(row = 5, column=0,padx = 5, pady = (0,5), sticky= "nsew")

 
mode_selector = ttk.Checkbutton(widgets_frame,text='Mode', style= "Switch", command=toggle_mode)
mode_selector.grid(row = 6, column=0,padx = 5, pady = 10, sticky= "nsew")

treeFrame = ttk.Frame(frame)
treeFrame.grid(row = 0, column = 1, pady =10)
treeScroll =ttk.Scrollbar(treeFrame)
treeScroll.pack(side = 'right', fill= 'y')

col_name = ("Name", "Age", "Subscription","Employment")
data_view = ttk.Treeview(treeFrame, show='headings', 
                         yscrollcommand=treeScroll.set,columns= col_name, height = 13)

data_view.column("Name", width= 100)
data_view.column("Age", width= 50)
data_view.column("Subscription", width=100)
data_view.column("Employment", width = 100)
data_view.pack()
treeScroll.config(command= data_view.yview)

def load_data():
    path = "/Users/zarnabkhan/Documents/MSc-BDS_QMUL/Self Learning Projects/TKINTER EXCEL APP/people.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active

    list_values = list(sheet.values)
    print(list_values)
    for col_name in list_values[0]:
        data_view.heading(col_name, text=col_name)

    for value_tuple in list_values[1:]:
        data_view.insert('', tk.END, values = value_tuple)

load_data()


def insert_row():
    name = name_entry.get()
    age = int(age_entry.get())
    subscription_status = status_entry.get()
    employement_status = "Employed" if bool.get() else "Unemployed"
    print(name, age, subscription_status, employement_status)

    path = "/Users/zarnabkhan/Documents/MSc-BDS_QMUL/Self Learning Projects/TKINTER EXCEL APP/people.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    row_values = [name, age, subscription_status, employement_status]
    sheet.append(row_values)
    workbook.save(path)

    data_view.insert('', tk.END, values = row_values)



root.mainloop()     # event loop to launch application

